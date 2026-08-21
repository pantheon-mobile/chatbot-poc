import csv
import io
import re
import time
from datetime import date, datetime, time as datetime_time

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import get_column_letter, range_boundaries


def _decimal_places(number_format: str) -> int:
    match = re.search(r"[0#]+\.([0#]+)", number_format)
    return len(match.group(1)) if match else 0


def format_excel_value(value, number_format: str = "General", data_type: str = "") -> str:
    """保存済みセル値を、代表的なExcel表示形式を保ちながら文字列化する。"""
    if data_type == "e":
        return f"ERROR: {value}"
    if value is None:
        return ""
    if isinstance(value, datetime):
        if any(token in number_format.lower() for token in ["h", "s"]):
            return value.strftime("%Y-%m-%d %H:%M:%S").rstrip("0").rstrip(":")
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime_time):
        return value.strftime("%H:%M:%S").rstrip("0").rstrip(":")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        fmt = number_format or "General"
        if "%" in fmt:
            places = _decimal_places(fmt)
            return f"{value * 100:.{places}f}%"
        # 00000のような書式はコードの先頭ゼロを保持する。
        integer_pattern = re.sub(r'"[^"]*"', "", fmt).split(";")[0]
        if isinstance(value, int) and re.fullmatch(r"0+", integer_pattern):
            return f"{value:0{len(integer_pattern)}d}"
        places = _decimal_places(fmt)
        use_grouping = "," in fmt
        if places:
            rendered = f"{value:,.{places}f}" if use_grouping else f"{value:.{places}f}"
        elif use_grouping:
            rendered = f"{value:,.0f}"
        else:
            rendered = str(value)
        currency = next((symbol for symbol in ["¥", "￥", "$", "€", "£"] if symbol in fmt), "")
        return f"{currency}{rendered}" if currency else rendered
    return str(value)


def _sheet_bounds(sheet) -> tuple[int, int, int, int, list[str]]:
    tables = list(sheet.tables.values())
    if tables:
        bounds = [range_boundaries(table.ref) for table in tables]
        return (
            min(item[0] for item in bounds), min(item[1] for item in bounds),
            max(item[2] for item in bounds), max(item[3] for item in bounds),
            [table.name for table in tables]
        )
    dimension = sheet.calculate_dimension()
    min_col, min_row, max_col, max_row = range_boundaries(dimension)
    return min_col, min_row, max_col, max_row, []


def _has_visible_content(sheet) -> bool:
    for row in sheet.iter_rows():
        for cell in row:
            if not isinstance(cell, MergedCell) and cell.value not in (None, ""):
                return True
    return False


def _cell_text(formula_cell, cached_cell) -> str:
    cached_value = cached_cell.value
    if formula_cell.data_type == "f" and cached_value is None:
        return f"FORMULA_WITHOUT_CACHED_VALUE: {formula_cell.value}"
    data_type = formula_cell.data_type if formula_cell.data_type == "e" else cached_cell.data_type
    return format_excel_value(cached_value, formula_cell.number_format, data_type)


def _markdown_table(rows: list[list[str]]) -> str:
    if not rows:
        return ""
    width = max(len(row) for row in rows)
    normalized = [row + [""] * (width - len(row)) for row in rows]
    escaped = [[value.replace("|", "\\|").replace("\r\n", "<br>").replace("\n", "<br>")
                for value in row] for row in normalized]
    return "\n".join([
        "| " + " | ".join(escaped[0]) + " |",
        "| " + " | ".join(["---"] * width) + " |",
        *("| " + " | ".join(row) + " |" for row in escaped[1:])
    ])


def parse_xlsx_workbook(xlsx_bytes: bytes, workbook_name: str, datasource_id: str) -> dict:
    """表示シートを1シート1CSV/Markdownへ変換する。数式の再計算は行わない。"""
    load_started = time.perf_counter()
    formula_book = load_workbook(io.BytesIO(xlsx_bytes), data_only=False, read_only=False)
    cached_book = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=False)
    workbook_load_ms = round((time.perf_counter() - load_started) * 1000, 1)
    parts = []
    visible_nonempty = [
        sheet for sheet in formula_book.worksheets
        if sheet.sheet_state == "visible" and _has_visible_content(sheet)
    ]
    sheet_count = len(visible_nonempty)
    for visible_index, formula_sheet in enumerate(visible_nonempty, start=1):
        parse_started = time.perf_counter()
        cached_sheet = cached_book[formula_sheet.title]
        min_col, min_row, max_col, max_row, table_names = _sheet_bounds(formula_sheet)
        rows = []
        comments = []
        for row_index in range(min_row, max_row + 1):
            row_values = []
            for column_index in range(min_col, max_col + 1):
                formula_cell = formula_sheet.cell(row=row_index, column=column_index)
                cached_cell = cached_sheet.cell(row=row_index, column=column_index)
                row_values.append(_cell_text(formula_cell, cached_cell))
                if formula_cell.comment and formula_cell.comment.text.strip():
                    comments.append(f"{formula_cell.coordinate}: {formula_cell.comment.text.strip()}")
            rows.append(row_values)
        sheet_parse_ms = round((time.perf_counter() - parse_started) * 1000, 1)

        csv_started = time.perf_counter()
        csv_buffer = io.StringIO(newline="")
        writer = csv.writer(csv_buffer, dialect="excel", lineterminator="\r\n")
        writer.writerows(rows)
        csv_bytes = ("\ufeff" + csv_buffer.getvalue()).encode("utf-8")
        csv_generation_ms = round((time.perf_counter() - csv_started) * 1000, 1)

        markdown_started = time.perf_counter()
        markdown_blocks = [f"# {workbook_name}", f"## シート: {formula_sheet.title}"]
        if table_names:
            markdown_blocks.append(f"### 表: {', '.join(table_names)}")
        markdown_blocks.append(_markdown_table(rows))
        if comments:
            markdown_blocks.extend(["### 注記", *(f"- {comment}" for comment in comments)])
        markdown_text = "\n\n".join(block for block in markdown_blocks if block).strip()
        markdown_generation_ms = round((time.perf_counter() - markdown_started) * 1000, 1)

        original_sheet_index = formula_book.sheetnames.index(formula_sheet.title) + 1
        document_part_id = f"{datasource_id}_SHEET_{visible_index:03d}"
        parts.append({
            "document_part_id": document_part_id,
            "sheet_name": formula_sheet.title,
            "sheet_index": original_sheet_index,
            "sheet_count": sheet_count,
            "cell_range": f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}",
            "table_name": ", ".join(table_names),
            "row_count": len(rows),
            "column_count": max((len(row) for row in rows), default=0),
            "sheet_parse_ms": sheet_parse_ms,
            "csv_generation_ms": csv_generation_ms,
            "markdown_generation_ms": markdown_generation_ms,
            "csv_bytes": csv_bytes,
            "markdown_text": markdown_text,
            "preview_rows": rows[:30]
        })
    formula_book.close()
    cached_book.close()
    generated_file_count = 1 + (2 * len(parts))
    generated_total_bytes = len(xlsx_bytes) + sum(
        len(part["csv_bytes"]) + len(part["markdown_text"].encode("utf-8")) for part in parts
    )
    return {
        "workbook_name": workbook_name, "workbook_load_ms": workbook_load_ms,
        "sheet_count": sheet_count, "parts": parts,
        "generated_file_count": generated_file_count,
        "generated_total_bytes": generated_total_bytes
    }
