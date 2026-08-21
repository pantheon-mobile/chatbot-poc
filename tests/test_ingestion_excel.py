from io import BytesIO

from openpyxl import Workbook
from openpyxl.worksheet.table import Table

from ingestion_excel import parse_xlsx_workbook


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "給付奨学金"
    sheet.append(["区分", "金額", "割合", "コード"])
    sheet.append(["第1区分", 38300, 0.25, 123])
    sheet["B2"].number_format = "¥#,##0"
    sheet["C2"].number_format = "0%"
    sheet["D2"].number_format = "00000"
    sheet.add_table(Table(displayName="支給月額", ref="A1:D2"))
    hidden = workbook.create_sheet("非表示")
    hidden["A1"] = "検索対象外"
    hidden.sheet_state = "hidden"
    empty = workbook.create_sheet("空シート")
    empty.sheet_state = "visible"
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_xlsx_conversion_excludes_hidden_and_empty_sheets():
    result = parse_xlsx_workbook(_workbook_bytes(), "sample.xlsx", "DS_TEST")

    assert result["sheet_count"] == 1
    assert result["parts"][0]["sheet_name"] == "給付奨学金"
    assert result["parts"][0]["document_part_id"] == "DS_TEST_SHEET_001"
    assert result["parts"][0]["cell_range"] == "A1:D2"
    assert result["parts"][0]["table_name"] == "支給月額"


def test_xlsx_conversion_preserves_common_display_formats_and_outputs():
    result = parse_xlsx_workbook(_workbook_bytes(), "sample.xlsx", "DS_TEST")
    part = result["parts"][0]

    assert part["preview_rows"][1] == ["第1区分", "¥38,300", "25%", "00123"]
    assert part["csv_bytes"].startswith(b"\xef\xbb\xbf")
    assert "# sample.xlsx" in part["markdown_text"]
    assert "## シート: 給付奨学金" in part["markdown_text"]
    assert "### 表: 支給月額" in part["markdown_text"]


def test_merged_cell_keeps_top_left_value():
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:B1")
    sheet["A1"] = "結合見出し"
    sheet["A2"] = "値"
    buffer = BytesIO()
    workbook.save(buffer)

    result = parse_xlsx_workbook(buffer.getvalue(), "merged.xlsx", "DS_MERGED")

    assert result["parts"][0]["preview_rows"][0] == ["結合見出し", ""]
